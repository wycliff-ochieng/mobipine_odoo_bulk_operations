# -*- coding: utf-8 -*-
import base64
import csv
import io
import json

from odoo import models, fields, api, _
from odoo.exceptions import UserError


class BulkOperationImportMapping(models.TransientModel):
    _name = 'bulk.operation.import.mapping'
    _description = 'Import Column Mapping'
    _order = 'column_index'

    wizard_id = fields.Many2one('bulk.operation.import.wizard', string='Wizard', required=True, ondelete='cascade')
    column_index = fields.Integer(string='Column #', required=True)
    column_name = fields.Char(string='Column Header')
    field_name = fields.Selection([
        ('__skip__', '— Do not import —'),
        ('distributor_id', 'Distributor ID'),
        ('customer_code', 'Customer ID'),
        ('customer_name', 'Customer Name'),
        ('customer_group', 'Customer Group'),
        ('branch_name', 'Branch / Location Name'),
        ('edition', 'Edition / Product Code'),
        ('date', 'Delivery Date'),
        ('delivered', 'Quantity Delivered'),
        ('returned', 'Quantity Returned'),
        ('price_unit', 'Unit Price'),
    ], string='Map to Field', required=True, default='__skip__')
    sample_values = fields.Char(string='Sample Values')
    is_imported = fields.Boolean(string='Import', default=True)


class BulkOperationImportWizard(models.TransientModel):
    _name = 'bulk.operation.import.wizard'
    _description = 'Bulk Operation Import Wizard'

    batch_id = fields.Many2one('bulk.operation.batch', string='Batch', required=True)

    file = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='Filename')
    file_type = fields.Char(string='File Type', compute='_compute_file_type')

    sheet_name = fields.Char(string='Sheet')
    sheet_options = fields.Char(string='Available Sheets')

    separator = fields.Selection([
        (',', 'Comma (,)'),
        (';', 'Semicolon (;)'),
        ('\t', 'Tab'),
    ], string='Separator', default=',')
    has_headers = fields.Boolean(string='File has headers', default=True)

    mapping_ids = fields.One2many('bulk.operation.import.mapping', 'wizard_id', string='Column Mapping')
    preview_data = fields.Text(string='Preview')

    state = fields.Selection([
        ('choose', 'Choose File'),
        ('mapping', 'Field Mapping'),
    ], string='Status', default='choose')

    def _compute_file_type(self):
        for wiz in self:
            if wiz.file_name:
                name = wiz.file_name.lower()
                if name.endswith('.xlsx') or name.endswith('.xls'):
                    wiz.file_type = 'xlsx'
                elif name.endswith('.csv'):
                    wiz.file_type = 'csv'
                else:
                    wiz.file_type = 'csv'
            else:
                wiz.file_type = False

    def action_parse_file(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_('Please select a file first.'))

        data = base64.b64decode(self.file)
        filename = (self.file_name or '').lower()

        if filename.endswith(('.xlsx', '.xls')):
            headers, sheets = self._parse_xlsx(data)
            if len(sheets) > 1 and not self.sheet_name:
                self.sheet_options = json.dumps(sheets)
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': 'bulk.operation.import.wizard',
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'new',
                    'context': {'sheet_selection': True},
                }
            rows = self._read_xlsx_sheet(data, self.sheet_name or sheets[0])
        else:
            headers, rows = self._parse_csv(data, self.separator or ',')

        self._create_mappings(headers, rows)
        self.state = 'mapping'

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'bulk.operation.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _parse_csv(self, data, separator=','):
        text = data.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(text), delimiter=separator)
        all_rows = list(reader)
        if not all_rows:
            raise UserError(_('The file appears to be empty.'))
        headers = [h.strip() for h in all_rows[0]] if self.has_headers else [f'Column {i+1}' for i in range(len(all_rows[0]))]
        rows = all_rows[1:] if self.has_headers else all_rows
        return headers, rows

    def _parse_xlsx(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('openpyxl is required to read Excel files.'))
        wb = load_workbook(io.BytesIO(data), read_only=True)
        sheets = wb.sheetnames
        headers = []
        if self.sheet_name or sheets:
            ws = wb[self.sheet_name or sheets[0]]
            for row in ws.iter_rows(values_only=True):
                headers = [str(h).strip() if h else '' for h in row]
                break
        wb.close()
        return headers, sheets

    def _read_xlsx_sheet(self, data, sheet_name):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('openpyxl is required to read Excel files.'))
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 and self.has_headers:
                continue
            if all(v in (None, '') for v in row):
                continue
            rows.append([str(v) if v is not None else '' for v in row])
        wb.close()
        return rows

    def _auto_match_field(self, header):
        header_lower = header.lower().strip().replace(' ', '_').replace('-', '_')
        field_map = {
            'distributor_id': ['distributor', 'distributorid', 'dist_id', 'route'],
            'customer_code': ['customerid', 'customer_id', 'cust_id', 'customer_code', 'code', 'ref'],
            'customer_name': ['name1', 'customer_name', 'customername', 'name', 'customer'],
            'customer_group': ['customergroup', 'customer_group', 'group', 'cust_group'],
            'branch_name': ['bulkname', 'branch_name', 'branch', 'location', 'branchname', 'route_location'],
            'edition': ['edition', 'product', 'product_code', 'item', 'sku', 'default_code'],
            'date': ['deliverydate', 'delivery_date', 'date', 'transaction_date'],
            'delivered': ['delivered', 'qty_delivered', 'quantity', 'qty'],
            'returned': ['returns', 'returned', 'return_qty'],
            'price_unit': ['price', 'price_unit', 'unit_price', 'selling_price'],
        }
        for field, aliases in field_map.items():
            if header_lower in aliases or header_lower == field:
                return field
        return '__skip__'

    def _create_mappings(self, headers, rows):
        self.mapping_ids.unlink()
        sample_count = min(3, len(rows))
        for idx, header in enumerate(headers):
            field = self._auto_match_field(header)
            samples = []
            for ri in range(sample_count):
                if ri < len(rows) and idx < len(rows[ri]):
                    samples.append(rows[ri][idx][:50])
                else:
                    samples.append('')
            sample_str = ', '.join(s for s in samples if s)
            self.env['bulk.operation.import.mapping'].create({
                'wizard_id': self.id,
                'column_index': idx,
                'column_name': header,
                'field_name': field,
                'sample_values': sample_str[:120],
                'is_imported': field != '__skip__',
            })

    def action_import(self):
        self.ensure_one()
        if not self.mapping_ids:
            raise UserError(_('No columns mapped for import.'))

        data = base64.b64decode(self.file)
        filename = (self.file_name or '').lower()

        if filename.endswith(('.xlsx', '.xls')):
            rows = self._read_xlsx_sheet(data, self.sheet_name or '')
        else:
            _, rows = self._parse_csv(data, self.separator or ',')

        active_mappings = self.mapping_ids.filtered(lambda m: m.is_imported and m.field_name != '__skip__')
        if not active_mappings:
            raise UserError(_('No columns selected for import.'))

        field_indices = {}
        for mapping in active_mappings:
            field_indices[mapping.field_name] = mapping.column_index

        lines_to_create = []
        skipped = []
        for index, row in enumerate(rows, start=1):
            vals = {}
            for field, col_idx in field_indices.items():
                if col_idx < len(row):
                    vals[field] = row[col_idx].strip() if row[col_idx] else ''
                else:
                    vals[field] = ''

            customer_code = vals.get('customer_code', '')
            customer_name = vals.get('customer_name', '')
            customer_group = vals.get('customer_group', '')
            branch_name = vals.get('branch_name', '')
            edition = vals.get('edition', '')
            distributor_id = vals.get('distributor_id', '')
            date_raw = vals.get('date', '')
            delivered_raw = vals.get('delivered', '0')
            returned_raw = vals.get('returned', '0')
            price_raw = vals.get('price_unit', '')

            if not date_raw:
                skipped.append(_("Row %s: missing delivery date") % index)
                continue

            line_date = self._parse_date(date_raw)
            if not line_date:
                skipped.append(_("Row %s: unreadable date '%s'") % (index, date_raw))
                continue

            try:
                delivered = float(delivered_raw or 0)
                returned = abs(float(returned_raw or 0))
            except (TypeError, ValueError):
                skipped.append(_("Row %s: 'delivered'/'returns' is not numeric") % index)
                continue

            try:
                price_unit = float(price_raw) if price_raw else 0.0
            except (TypeError, ValueError):
                price_unit = 0.0

            partner = self.batch_id._resolve_partner(customer_code, customer_name)
            product = self.batch_id._resolve_product(edition)
            location = self.batch_id._resolve_location(branch_name, partner)

            errors = []
            if not partner:
                errors.append(_("customer '%s' not found") % (customer_name or customer_code))
            if not product:
                errors.append(_("product/edition '%s' not found") % edition)
            if not location:
                msg = _("no stock location found")
                if branch_name:
                    msg = _("location '%s' not found — check that the name matches an internal stock location in Odoo") % branch_name
                errors.append(msg)

            if not price_unit and product:
                price_unit = product.lst_price

            lines_to_create.append({
                'batch_id': self.batch_id.id,
                'distributor_id': distributor_id,
                'customer_code': customer_code,
                'customer_name': customer_name,
                'customer_group': customer_group,
                'branch_name': branch_name,
                'edition': edition,
                'partner_id': partner.id if partner else False,
                'product_id': product.id if product else False,
                'location_id': location.id if location else False,
                'date': line_date,
                'delivered': delivered,
                'returned': returned,
                'price_unit': price_unit,
                'error_message': '; '.join(errors) if errors else False,
            })

        if lines_to_create:
            self.env['bulk.operation.line'].create(lines_to_create)
            self.batch_id.write({'state': 'imported'})

        unresolved = len([l for l in lines_to_create if l.get('error_message')])
        msg = _("Imported %s row(s).") % len(lines_to_create)
        if unresolved:
            msg += "<br/>" + _("%s row(s) need attention - see the Error column.") % unresolved
        if skipped:
            rpt = ["<br/>" + _("Row %s could not be read.") % s for s in skipped[:30]]
            msg += "<br/>" + _("%s row(s) could not be read:") % len(skipped) + "".join(rpt[:30])
            if len(skipped) > 30:
                msg += "<br/>" + _("... and %s more.") % (len(skipped) - 30)
        self.batch_id.message_post(body=msg)

        return {
            'type': 'ir.actions.act_window_close',
        }

    @api.model
    def _parse_date(self, date_raw):
        from datetime import datetime, date
        if isinstance(date_raw, datetime):
            return date_raw.date()
        if isinstance(date_raw, date):
            return date_raw
        try:
            from odoo import fields as odoo_fields
            return odoo_fields.Date.to_date(str(date_raw).strip())
        except Exception:
            return False
