# -*- coding: utf-8 -*-
import base64
import csv
import io
import logging
import threading
from datetime import date, datetime, timedelta

import odoo
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class BulkOperationBatch(models.Model):
    _name = 'bulk.operation.batch'
    _description = 'Bulk Operation Batch'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'id desc'

    @api.model
    def load(self, fields, data):
        raise UserError(_(
            "Direct CSV/Excel import into Batches is not supported. "
            "Upload your file via the \"Import File\" binary field on the batch form, "
            "then click \"Process File\" to use the Quick Import, "
            "or click \"Standard Import\" for the column-mapping wizard."
        ))

    # Maps the source file's header names (lower-case, exactly as they
    # appear in the export) to the internal keys used while importing.
    # Edit this single dict if the source file's headers ever change -
    # nothing else in the import logic needs to change.
    COLUMN_MAP = {
        'distributorid': 'distributor_id',
        'customerid': 'customer_code',
        'name1': 'customer_name',
        'customergroup': 'customer_group',
        'bulkname': 'branch_name',
        'edition': 'edition',
        'deliverydate': 'date',
        'delivered': 'delivered',
        'returns': 'returned',
    }

    name = fields.Char(string='Batch Reference', required=True, copy=False,
                        readonly=True, default=lambda self: _('New'))
    import_mode = fields.Selection([
        ('quick', 'Quick Import'),
        ('standard', 'Standard Import'),
    ], string='Import Method', default='quick', required=True,
        help="Quick: upload and auto-import with predefined column mapping.\n"
             "Standard: choose fields to import, map columns, multi-sheet support.")
    state = fields.Selection([
        ('draft', 'Draft'),
        ('imported', 'Imported'),
        ('processed', 'Processed'),
    ], string='Status', default='draft', readonly=True, tracking=True, copy=False)

    import_file = fields.Binary(string='Upload File (CSV or XLSX)', attachment=True)
    filename = fields.Char(string='Filename')

    processing_progress = fields.Float(string='Progress', readonly=True, default=0.0,
        help="Processing progress (0-100). Updated periodically during background processing.")
    processing_status = fields.Char(string='Status', readonly=True, default='',
        help="Current status message shown during background processing.")

    line_ids = fields.One2many('bulk.operation.line', 'batch_id', string='Imported Lines')
    error_line_count = fields.Integer(compute='_compute_line_counts')
    pending_line_count = fields.Integer(compute='_compute_line_counts')

    sale_order_ids = fields.Many2many('sale.order', string='Generated Sales Orders', readonly=True, copy=False)
    purchase_order_ids = fields.Many2many('purchase.order', string='Generated Purchase Orders', readonly=True, copy=False)
    invoice_ids = fields.Many2many('account.move', string='Generated Invoices', readonly=True, copy=False)
    picking_ids = fields.Many2many('stock.picking', compute='_compute_picking_ids', string='Generated Deliveries')

    @api.depends('line_ids.error_message', 'line_ids.is_processed')
    def _compute_line_counts(self):
        for batch in self:
            batch.error_line_count = len(batch.line_ids.filtered(lambda l: l.error_message))
            batch.pending_line_count = len(batch.line_ids.filtered(lambda l: not l.is_processed and not l.error_message))

    @api.depends('sale_order_ids')
    def _compute_picking_ids(self):
        for batch in self:
            batch.picking_ids = batch.sale_order_ids.picking_ids

    def action_open_import_wizard(self):
        self.ensure_one()
        return {
            'name': _('Standard Import'),
            'type': 'ir.actions.act_window',
            'res_model': 'bulk.operation.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_batch_id': self.id},
        }

    def action_view_lines(self):
        """Open the full-screen imported lines list, filtered to this batch,
        grouped by date by default so the user sees one date block at a time."""
        self.ensure_one()
        return {
            'name': _('Imported Data — %s') % self.name,
            'type': 'ir.actions.act_window',
            'res_model': 'bulk.operation.line',
            'view_mode': 'list',
            'domain': [('batch_id', '=', self.id)],
            'context': {
                'default_batch_id': self.id,
                'search_default_group_date': 1,
            },
        }

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('bulk.operation.batch') or _('New')
        return super().create(vals_list)

    # ------------------------------------------------------------------
    # File reading (CSV or XLSX - detected automatically)
    # ------------------------------------------------------------------
    def _read_rows(self, data):
        """Returns a list of dicts: one per data row, keyed by lower-cased,
        stripped header name - regardless of whether the source was CSV or
        Excel."""
        filename = (self.filename or '').lower()
        is_excel = filename.endswith(('.xlsx', '.xls')) or data[:2] == b'PK'
        if is_excel:
            return self._read_rows_xlsx(data)
        return self._read_rows_csv(data)

    def _read_rows_xlsx(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_(
                "This server cannot read Excel files (the 'openpyxl' library is missing). "
                "Please export the file as CSV instead, or ask your administrator to install openpyxl."
            ))
        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheet = workbook.worksheets[0]
            rows_iter = sheet.iter_rows(values_only=True)
            header = [str(h).strip().lower() if h is not None else '' for h in next(rows_iter)]
        except Exception as e:
            raise UserError(_("Failed to read the Excel file: %s") % str(e))

        rows = []
        for raw_row in rows_iter:
            if all(v in (None, '') for v in raw_row):
                continue
            rows.append(dict(zip(header, raw_row)))
        return rows

    def _read_rows_csv(self, data):
        text = None
        last_err = None
        for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                text = data.decode(enc)
                break
            except UnicodeDecodeError as e:
                last_err = e
        if text is None:
            raise UserError(_(
                "Failed to read the file as CSV (tried utf-8 and common Windows/Excel encodings). "
                "If this is really an Excel file, make sure the filename ends in .xlsx. Error: %s"
            ) % last_err)

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for raw_row in reader:
            rows.append({(k or '').strip().lower(): v for k, v in raw_row.items()})
        return rows

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------
    def action_import_file(self):
        """Parse the uploaded file using COLUMN_MAP, resolve each row's
        Customer / Product / Location, and stage everything as review-able
        lines - exactly mirroring the source columns plus the resolved
        links. Rows that can't be fully resolved are imported anyway and
        flagged with an error_message, so they can be fixed by hand in the
        list before processing."""
        self.ensure_one()
        if not self.import_file:
            raise UserError(_("Please upload a file first."))

        self.line_ids.unlink()
        data = base64.b64decode(self.import_file)
        rows = self._read_rows(data)
        if not rows:
            raise UserError(_("The file appears to be empty."))

        # ---- Phase 1: collect unique resolution keys ----
        editions = set()
        code_to_name = {}
        names = set()
        branch_names = set()

        for raw_row in rows:
            vals = {field: raw_row.get(src) for src, field in self.COLUMN_MAP.items()}
            edition = str(vals.get('edition') or '').strip()
            customer_code = str(vals.get('customer_code') or '').strip()
            customer_name = str(vals.get('customer_name') or '').strip()
            branch_name = str(vals.get('branch_name') or '').strip()
            if edition:
                editions.add(edition)
            if customer_code:
                code_to_name[customer_code] = customer_name
            if customer_name:
                names.add(customer_name)
            if branch_name:
                branch_names.add(branch_name)

        # ---- Phase 2: batch-resolve all unique entities (handful of queries) ----
        product_map = self._batch_resolve_products(editions)
        partner_map = self._batch_resolve_partners(code_to_name, names)
        location_map = self._batch_resolve_locations(branch_names)

        # ---- Phase 3: process each row using pre-resolved maps ----
        lines_to_create = []
        skipped = []
        for index, raw_row in enumerate(rows, start=2):  # row 1 is the header
            vals = {field: raw_row.get(src) for src, field in self.COLUMN_MAP.items()}

            customer_code = str(vals.get('customer_code') or '').strip()
            customer_name = str(vals.get('customer_name') or '').strip()
            customer_group = str(vals.get('customer_group') or '').strip()
            branch_name = str(vals.get('branch_name') or '').strip()
            edition = str(vals.get('edition') or '').strip()
            distributor_id = str(vals.get('distributor_id') or '').strip()
            date_raw = vals.get('date')

            if not date_raw:
                skipped.append(_("Row %s: missing delivery date") % index)
                continue
            line_date = self._parse_cell_date(date_raw)
            if not line_date:
                skipped.append(_("Row %s: unreadable date '%s'") % (index, date_raw))
                continue

            try:
                delivered = float(vals.get('delivered') or 0)
                returned = abs(float(vals.get('returned') or 0))
            except (TypeError, ValueError):
                skipped.append(_("Row %s: 'delivered'/'returns' is not numeric") % index)
                continue

            partner = partner_map.get(customer_code) or partner_map.get('name:' + customer_name)
            product = product_map.get(edition)
            location = location_map.get(branch_name)

            errors = []
            if not partner:
                errors.append(_("customer '%s' not found") % (customer_name or customer_code))
            if not product:
                errors.append(_("product/edition '%s' not found") % edition)
            if not location:
                msg = _("no stock location found")
                if branch_name:
                    msg = _("location '%s' not found — check that the name matches an internal stock location in Odoo") % branch_name
                errors.append(msg)

            lines_to_create.append({
                'batch_id': self.id,
                'distributor_id': distributor_id,
                'customer_code': customer_code,
                'customer_name': customer_name,
                'customer_group': customer_group,
                'branch_name': branch_name,
                'edition': edition,
                'partner_id': partner.id if partner else False,
                'product_id': product.id if product else False,
                'location_id': location.id if location else False,
                'date': line_date,
                'delivered': delivered,
                'returned': returned,
                'price_unit': product.lst_price if product else 0.0,
                'error_message': '; '.join(errors) if errors else False,
            })

        if lines_to_create:
            self.env['bulk.operation.line'].create(lines_to_create)
            self.write({
                'state': 'imported',
                'processing_progress': 0.0,
                'processing_status': '',
            })

        unresolved = len([l for l in lines_to_create if l.get('error_message')])
        msg = _("Imported %s row(s).") % len(lines_to_create)
        if unresolved:
            msg += "<br/>" + _("%s row(s) need attention - see the Error column.") % unresolved
        if skipped:
            msg += "<br/>" + _("%s row(s) could not be read at all:") % len(skipped) + "<br/>" + "<br/>".join(skipped[:30])
            if len(skipped) > 30:
                msg += "<br/>" + _("... and %s more.") % (len(skipped) - 30)
        self.message_post(body=msg)

        if not lines_to_create:
            raise UserError(_("No rows could be parsed from the file. Check the chatter for details."))

    @api.model
    def _parse_cell_date(self, date_raw):
        if isinstance(date_raw, datetime):
            return date_raw.date()
        if isinstance(date_raw, date):
            return date_raw
        try:
            return fields.Date.to_date(str(date_raw).strip())
        except Exception:
            return False

    def _resolve_partner(self, customer_code, customer_name):
        Partner = self.env['res.partner']
        partner = False
        if customer_code:
            partner = Partner.search([('ref', '=', customer_code)], limit=1)
        if not partner and customer_name:
            partner = Partner.search([('name', '=', customer_name)], limit=1)
            if not partner:
                partner = Partner.search([('name', '=ilike', customer_name)], limit=1)
        return partner

    def _resolve_product(self, edition):
        if not edition:
            return False
        Product = self.env['product.product']
        product = Product.search([('default_code', '=', edition)], limit=1)
        if not product:
            product = Product.search([('default_code', '=ilike', edition)], limit=1)
        if not product:
            product = Product.search([('name', '=', edition)], limit=1)
        if not product:
            product = Product.search([('name', '=ilike', edition)], limit=1)
        if not product:
            product = Product.search([('barcode', '=', edition)], limit=1)
        return product

    def _resolve_location(self, branch_name, partner):
        Location = self.env['stock.location']
        if branch_name:
            location = Location.search([('name', '=ilike', branch_name), ('usage', '=', 'internal')], limit=1)
            if location:
                return location
            location = Location.search([('complete_name', '=ilike', branch_name), ('usage', '=', 'internal')], limit=1)
            if location:
                return location
        return False

    # ------------------------------------------------------------------
    # Batch resolution helpers (minimise DB round-trips for large files)
    # ------------------------------------------------------------------
    def _batch_resolve_products(self, edition_values):
        """Return {edition: product} for every unique edition in the file.
        Uses batch ``in`` queries for exact matches so that thousands of
        rows resolve with a handful of SQL round-trips instead of N×5."""
        Product = self.env['product.product']
        result = {}
        remaining = set(e for e in edition_values if e)
        if not remaining:
            return result

        # Priority 1 — exact default_code matches (batched)
        found = Product.search([('default_code', 'in', list(remaining))])
        for p in found:
            if p.default_code in remaining:
                result[p.default_code] = p
        remaining.difference_update(result.keys())
        if not remaining:
            return result

        # Priority 2-5 — fallback cascade for the few remaining items
        for edition in list(remaining):
            product = Product.search([
                '|', '|', '|',
                ('default_code', '=ilike', edition),
                ('name', '=', edition),
                ('name', '=ilike', edition),
                ('barcode', '=', edition),
            ], limit=1)
            if product:
                result[edition] = product

        return result

    def _batch_resolve_partners(self, code_to_name, name_values):
        """Return {customer_code: partner, 'name:customer_name': partner}
        for every customer in the file."""
        Partner = self.env['res.partner']
        result = {}

        codes = [c for c in code_to_name if c]
        if codes:
            found = Partner.search([('ref', 'in', codes)])
            for p in found:
                if p.ref in code_to_name:
                    result[p.ref] = p

        names = [n for n in name_values if n]
        if names:
            found = Partner.search([('name', 'in', names)])
            for p in found:
                key = 'name:' + p.name
                if key not in result:
                    result[key] = p

        for name in names:
            key = 'name:' + name
            if key not in result:
                p = Partner.search([('name', '=ilike', name)], limit=1)
                if p:
                    result[key] = p

        return result

    def _batch_resolve_locations(self, branch_names):
        """Return {branch_name: stock.location} for every branch in file."""
        Location = self.env['stock.location']
        result = {}
        names = [n for n in branch_names if n]
        if not names:
            return result

        for name in names:
            loc = Location.search([
                ('name', '=ilike', name), ('usage', '=', 'internal'),
            ], limit=1)
            if not loc:
                loc = Location.search([
                    ('complete_name', '=ilike', name), ('usage', '=', 'internal'),
                ], limit=1)
            if loc:
                result[name] = loc

        return result

    # ------------------------------------------------------------------
    # Processing engine
    # ------------------------------------------------------------------
    def _get_effective_date(self, line_date):
        """Saturday/Sunday entries are folded into the following Monday,
        so a customer never gets two separate weekend Sales Orders."""
        weekday = line_date.weekday()
        if weekday == 5:      # Saturday
            return line_date + timedelta(days=2)
        elif weekday == 6:    # Sunday
            return line_date + timedelta(days=1)
        return line_date

    def action_batch_process(self):
        """Validate lines and start background processing.

        Processing is moved to a background thread so large imports don't
        hit the HTTP request timeout (default: 120 s).  The chatter on
        the batch will be updated when processing finishes.

        Each (date, customer) group runs in its own DB savepoint, so one
        failure doesn't block the rest of the batch.
        """
        self.ensure_one()
        if not self.line_ids.filtered(lambda l: not l.is_processed):
            raise UserError(_("No lines are ready to process."))

        self.message_post(body=_("Processing started in background…"))
        threading.Thread(
            target=self._process_in_background,
            args=(self.id, self._cr.dbname),
            daemon=True,
        ).start()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Processing Started'),
                'message': _(
                    'Batch processing started in the background. '
                    'Refresh the page and check the chatter for results.'
                ),
                'sticky': True,
            },
        }

    @api.model
    def _process_in_background(self, batch_id, db_name):
        """Entry point for the background thread: open a dedicated DB
        cursor + environment and run the full processing workflow."""
        try:
            with odoo.sql_db.db_connect(db_name).cursor() as cr:
                env = api.Environment(cr, odoo.SUPERUSER_ID, {})
                batch = env['bulk.operation.batch'].browse(batch_id)
                batch._execute_batch_process()
                cr.commit()
        except Exception:
            _logger.exception("Background batch #%s processing crashed", batch_id)
            try:
                with odoo.sql_db.db_connect(db_name).cursor() as cr:
                    env = api.Environment(cr, odoo.SUPERUSER_ID, {})
                    batch = env['bulk.operation.batch'].browse(batch_id)
                    batch.message_post(
                        body=_("Batch processing crashed unexpectedly. Check the server logs."))
                    cr.commit()
            except Exception:
                pass

    def _execute_batch_process(self):
        """Core processing: group lines by (date, partner), then process
        in manageable chunks, committing progress periodically so the
        user can see status by refreshing the form."""
        self.ensure_one()
        batch_id = self.id

        candidate_lines = self.line_ids.filtered(lambda l: not l.is_processed)
        if not candidate_lines:
            raise UserError(_("No lines are ready to process."))

        invalid = candidate_lines.filtered(
            lambda l: not (l.partner_id and l.product_id and l.location_id) or l.returned > l.delivered
        )
        for line in invalid:
            reasons = []
            if not line.partner_id:
                reasons.append(_("missing customer"))
            if not line.product_id:
                reasons.append(_("missing product"))
            if not line.location_id:
                reasons.append(_("missing location"))
            if line.returned > line.delivered:
                reasons.append(_("returned exceeds delivered"))
            line.error_message = '; '.join(reasons)

        valid_lines = candidate_lines - invalid
        if not valid_lines:
            raise UserError(_("No lines are ready to process. Check the Error column on the imported lines."))

        by_date = {}
        for line in valid_lines:
            eff_date = self._get_effective_date(line.date)
            by_date.setdefault(eff_date, {})
            by_date[eff_date].setdefault(line.partner_id, self.env['bulk.operation.line'])
            by_date[eff_date][line.partner_id] |= line

        groups = []
        for eff_date in sorted(by_date.keys()):
            for partner, lines in by_date[eff_date].items():
                groups.append((eff_date, partner, lines))

        total = len(groups)
        if not total:
            return

        self.write({
            'state': 'imported',
            'processing_progress': 0.0,
            'processing_status': _('Starting — %s group(s) to process') % total,
        })
        self.env.cr.commit()

        chunk_size = 15

        wh = self._get_warehouse()
        journal = self._get_payment_journal()
        payment_method = journal.inbound_payment_method_line_ids[:1] if journal else False

        all_sale_order_ids = []
        all_invoice_ids = []
        net_demand = {}
        errors = []
        processed_count = 0

        for idx, (eff_date, partner, lines) in enumerate(groups, start=1):
            try:
                with self.env.cr.savepoint():
                    sale_order = self._create_sale_order(partner, eff_date, lines, warehouse=wh)
                    self._process_delivery(sale_order, lines, warehouse=wh)
                    self._process_returns(sale_order, lines)
                    invoice = self._process_invoice(
                        sale_order, journal=journal, payment_method=payment_method)

                    for line in lines:
                        if line.net_qty > 0:
                            net_demand[line.product_id] = net_demand.get(line.product_id, 0.0) + line.net_qty

                    lines.write({'is_processed': True, 'error_message': False})
                    all_sale_order_ids.append(sale_order.id)
                    if invoice:
                        all_invoice_ids.append(invoice.id)
                    processed_count += 1
            except Exception as e:
                lines.write({'error_message': str(e)[:500]})
                errors.append(_("%(partner)s on %(date)s: %(error)s") % {
                    'partner': partner.name, 'date': eff_date, 'error': str(e),
                })

            if idx % chunk_size == 0 or idx == total:
                progress = min(98.0, (idx / total) * 100.0)
                status = _('Processed %(done)s/%(total)s groups') % {'done': idx, 'total': total}
                self.write({
                    'processing_progress': progress,
                    'processing_status': status,
                })
                self.env.cr.commit()

        self.write({
            'processing_progress': 99.0,
            'processing_status': _('Creating purchase orders…'),
        })
        self.env.cr.commit()

        purchase_orders = self._create_consolidated_po(net_demand)

        self.write({
            'processing_progress': 100.0,
            'processing_status': _('Completed'),
            'state': 'processed',
            'sale_order_ids': [(4, so_id) for so_id in all_sale_order_ids],
            'invoice_ids': [(4, inv_id) for inv_id in all_invoice_ids],
            'purchase_order_ids': [(4, po.id) for po in purchase_orders],
        })
        self.env.cr.commit()

        summary = _("Batch processed: %(ok)s group(s) succeeded, %(err)s failed.") % {
            'ok': processed_count, 'err': len(errors),
        }
        if errors:
            summary += "<br/>" + "<br/>".join(errors)
        self.message_post(body=summary)
        self.env.cr.commit()

    def _get_warehouse(self):
        return self.env['stock.warehouse'].search([
            ('company_id', '=', self.env.company.id),
        ], limit=1)

    def _get_payment_journal(self):
        journal = self.env['account.journal'].search([
            ('type', '=', 'cash'), ('company_id', '=', self.env.company.id),
        ], limit=1)
        if not journal:
            journal = self.env['account.journal'].search([
                ('type', '=', 'bank'), ('company_id', '=', self.env.company.id),
            ], limit=1)
        return journal

    def _create_sale_order(self, partner, eff_date, lines, warehouse=None):
        wh = warehouse or self._get_warehouse()
        sale_order = self.env['sale.order'].create({
            'partner_id': partner.id,
            'date_order': fields.Datetime.to_datetime(eff_date),
            'origin': self.name,
            'warehouse_id': wh.id if wh else False,
            'order_line': [(0, 0, {
                'product_id': line.product_id.id,
                'product_uom_qty': line.delivered,
                'price_unit': line.price_unit,
            }) for line in lines],
        })
        for so_line, line in zip(sale_order.order_line, lines):
            line.sale_line_id = so_line.id
        sale_order.action_confirm()
        return sale_order

    def _process_delivery(self, sale_order, lines, warehouse=None):
        location_by_product = {line.product_id.id: line.location_id for line in lines}
        primary_location = lines[:1].location_id

        for picking in sale_order.picking_ids.filtered(lambda p: p.state not in ('done', 'cancel')):
            picking.location_id = primary_location.id
            for move in picking.move_ids:
                move.location_id = location_by_product.get(move.product_id.id, primary_location).id
                move.quantity = move.product_uom_qty
            picking.action_assign()
            insufficient = picking.move_ids.filtered(
                lambda m: m.product_uom_qty > 0 and m.quantity <= 0
            )
            if insufficient:
                wh = warehouse or self._get_warehouse()
                if wh and wh.lot_stock_id:
                    picking.location_id = wh.lot_stock_id.id
                    for move in picking.move_ids:
                        move.location_id = wh.lot_stock_id.id
                        move.quantity = move.product_uom_qty
                    picking.action_assign()
            result = picking.with_context(
                skip_backorder=True, skip_sms=True,
            ).button_validate()
            if isinstance(result, dict):
                raise UserError(_(
                    "Delivery could not be validated automatically "
                    "(likely insufficient stock at %s)."
                ) % primary_location.display_name)

    def _process_returns(self, sale_order, lines):
        for line in lines:
            if line.returned > 0:
                self._process_return(sale_order, line)

    def _process_return(self, sale_order, line):
        validated_picking = sale_order.picking_ids.filtered(
            lambda p: p.state == 'done' and p.picking_type_code == 'outgoing'
        )[:1]
        if not validated_picking:
            return

        picking_type = validated_picking.picking_type_id
        return_picking_type = self.env['stock.picking.type'].search([
            ('code', '=', 'incoming'),
            ('warehouse_id', '=', picking_type.warehouse_id.id),
        ], limit=1) or self.env['stock.picking.type'].search([('code', '=', 'incoming')], limit=1)

        return_picking = self.env['stock.picking'].create({
            'partner_id': sale_order.partner_id.id,
            'picking_type_id': return_picking_type.id,
            'location_id': validated_picking.location_dest_id.id,
            'location_dest_id': line.location_id.id or validated_picking.location_id.id,
            'origin': _("Return: %s") % sale_order.name,
        })

        self.env['stock.move'].create({
            'name': line.product_id.name,
            'product_id': line.product_id.id,
            'product_uom_qty': line.returned,
            'product_uom': line.product_id.uom_id.id,
            'picking_id': return_picking.id,
            'location_id': return_picking.location_id.id,
            'location_dest_id': return_picking.location_dest_id.id,
        })

        return_picking.action_confirm()
        return_picking.action_assign()
        for move in return_picking.move_ids:
            move.quantity = move.product_uom_qty
        return_picking.with_context(skip_backorder=True, skip_sms=True).button_validate()

    def _process_invoice(self, sale_order, journal=None, payment_method=None):
        invoice = sale_order._create_invoices()
        if not invoice:
            return invoice
        invoice.action_post()
        self._register_payment(invoice, journal=journal, payment_method=payment_method)
        return invoice

    def _register_payment(self, invoice, journal=None, payment_method=None):
        journal = journal or self._get_payment_journal()
        if not journal:
            self.message_post(body=_("No cash or bank journal found - invoice %s left unpaid.") % invoice.name)
            return

        payment_method = payment_method or journal.inbound_payment_method_line_ids[:1]
        if not payment_method:
            self.message_post(body=_("Journal %s has no inbound payment method - invoice %s left unpaid.") % (
                journal.name, invoice.name))
            return

        payment_register = self.env['account.payment.register'].with_context(
            active_model='account.move', active_ids=invoice.ids,
        ).create({
            'journal_id': journal.id,
            'payment_method_line_id': payment_method.id,
            'amount': invoice.amount_total,
        })
        payment_register.action_create_payments()

    # ------------------------------------------------------------------
    # Purchasing
    # ------------------------------------------------------------------
    def _get_fallback_supplier(self):
        param = self.env['ir.config_parameter'].sudo()
        supplier_name = param.get_param('bulk_operations.default_supplier', 'Deen Innovations')
        supplier = self.env['res.partner'].search([('name', '=', supplier_name)], limit=1)
        if not supplier:
            supplier = self.env['res.partner'].create({'name': supplier_name, 'supplier_rank': 1})
        return supplier

    def _create_consolidated_po(self, net_demand):
        """One consolidated Purchase Order per supplier: each product's own
        vendor (product.seller_ids) is used where configured, falling back
        to a single default supplier otherwise."""
        if not net_demand:
            return self.env['purchase.order']

        fallback_supplier = self._get_fallback_supplier()
        grouped_by_supplier = {}
        for product, qty in net_demand.items():
            supplier = product.seller_ids[:1].partner_id if product.seller_ids else False
            supplier = supplier or fallback_supplier
            grouped_by_supplier.setdefault(supplier, []).append((product, qty))

        purchase_orders = self.env['purchase.order']
        for supplier, product_qty_list in grouped_by_supplier.items():
            po = self.env['purchase.order'].create({
                'partner_id': supplier.id,
                'date_order': fields.Datetime.now(),
                'origin': self.name,
            })
            for product, qty in product_qty_list:
                seller = product.seller_ids.filtered(lambda s: s.partner_id == supplier)[:1]
                price = seller.price if seller else (product.standard_price or 1.0)
                self.env['purchase.order.line'].create({
                    'order_id': po.id,
                    'product_id': product.id,
                    'name': product.name,
                    'product_qty': qty,
                    'product_uom': product.uom_id.id,
                    'price_unit': price,
                    'date_planned': fields.Datetime.now(),
                })
            purchase_orders |= po
        return purchase_orders
